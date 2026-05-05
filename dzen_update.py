"""
Дзен Аналитика — скрипт обновления данных
Использование:  python dzen_update.py '<json_string>'
JSON-структура:  {"pubs": [...], "dailyIntervals": [...], "subIntervals": [...], "channelTotal": {...}}
"""

import json
import sys
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

# ─── Пути ────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH  = os.path.join(SCRIPT_DIR, "dzen_analytics.xlsx")

# ─── Цвета / стили ───────────────────────────────────────────────────────────
HEADER_BG   = "1A1A2E"   # тёмно-синий
HEADER_FG   = "FFFFFF"
TOTAL_BG    = "16213E"
TOTAL_FG    = "F5A623"
ALT_ROW_BG  = "F0F4FF"
ACCENT      = "0F3460"
GREEN_BG    = "E8F5E9"
SNAP_BG     = "FFF9C4"   # жёлтый — дата снятия

def thin_border():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)

def header_style(cell, bg=HEADER_BG, fg=HEADER_FG, size=10, bold=True):
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    cell.border = thin_border()

def data_style(cell, bg=None, bold=False, align="right", fmt=None):
    cell.font = Font(name="Arial", size=9, bold=bold)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = thin_border()
    if fmt:
        cell.number_format = fmt

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


# ═══════════════════════════════════════════════════════════════════════════════
# Лист 1 — Публикации (накапливающиеся снапшоты)
# ═══════════════════════════════════════════════════════════════════════════════
PUB_HEADERS = [
    "Дата снятия",
    "Название публикации",
    "ID публикации",
    "Тип",
    "Дата публикации",
    "Показы",
    "Открытия",
    "Дочитывания",
    "Просмотры видео/гифок",
    "Подписки",
    "Лайки",
    "Комментарии",
    "Репосты",
    "Время просмотра, мин",
    "CTR, %",
    "% дочитываний",
]

PUB_WIDTHS = [13, 55, 26, 10, 16, 9, 9, 13, 18, 10, 8, 13, 9, 20, 8, 15]

PUB_TYPE_RU = {
    "article": "Статья",
    "gif":     "Пост/Гифка",
    "video":   "Видео",
    "brief":   "Пост",
    "post":    "Пост",
}

def write_publications_sheet(ws, pubs_rows, snapshot_date, is_new=True):
    """Добавляет строки снапшота на лист Публикации."""
    if is_new:
        # Заголовки
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 36
        for ci, h in enumerate(PUB_HEADERS, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            header_style(cell)
        for ci, w in enumerate(PUB_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    start_row = ws.max_row + 1 if not is_new else 2

    for ri, pub in enumerate(pubs_rows):
        row_bg = ALT_ROW_BG if ri % 2 == 0 else None
        r = start_row + ri

        vals = [
            snapshot_date,
            pub["title"],
            pub["id"],
            PUB_TYPE_RU.get(pub["type"], pub["type"]),
            pub["pub_date"],
            pub["impressions"],
            pub["page_views"],
            pub["deep_views"],
            pub["type_specific_views"],
            pub["subscriptions"],
            pub["likes"],
            pub["comments"],
            pub["shares"],
            pub["view_time_min"],
            pub["ctr_pct"],
            pub["deep_views_rate_pct"],
        ]
        aligns = ["center","left","left","center","center",
                  "right","right","right","right","right",
                  "right","right","right","right","right","right"]
        fmts   = [None, None, None, None, None,
                  "#,##0","#,##0","#,##0","#,##0","#,##0",
                  "#,##0","#,##0","#,##0","#,##0.0","0.00","0.00"]

        for ci, (v, al, fm) in enumerate(zip(vals, aligns, fmts), 1):
            cell = ws.cell(row=r, column=ci, value=v)
            bg = SNAP_BG if ci == 1 else row_bg
            data_style(cell, bg=bg, align=al, fmt=fm)
            if ci == 1:
                cell.font = Font(name="Arial", size=9, bold=True, color="7B6000")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(PUB_HEADERS))}1"


# ═══════════════════════════════════════════════════════════════════════════════
# Лист 2 — Канал по дням
# ═══════════════════════════════════════════════════════════════════════════════
DAILY_HEADERS = [
    "Дата",
    "Показы",
    "Открытия",
    "Дочитывания",
    "Просмотры видео",
    "Подписки",
    "Лайки",
    "Время просмотра, мин",
]
DAILY_WIDTHS = [13, 9, 9, 13, 16, 10, 8, 20]

def write_daily_sheet(ws, intervals, is_new=True):
    if is_new:
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 36
        for ci, h in enumerate(DAILY_HEADERS, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            header_style(cell)
        for ci, w in enumerate(DAILY_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    # Collect existing dates to avoid duplicates
    existing_dates = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            existing_dates.add(str(row[0]))

    added = 0
    for item in intervals:
        if str(item["date"]) in existing_dates:
            continue
        ri = ws.max_row + 1
        bg = ALT_ROW_BG if (ri % 2 == 0) else None
        vals = [item["date"], item["impressions"], item["page_views"],
                item["deep_views"], item["type_views"], item["subscriptions"],
                item["likes"], item["view_time_min"]]
        fmts = [None,"#,##0","#,##0","#,##0","#,##0","#,##0","#,##0","#,##0.0"]
        for ci, (v, fm) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            data_style(cell, bg=bg, align="center" if ci==1 else "right", fmt=fm)
        added += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(DAILY_HEADERS))}1"
    return added


# ═══════════════════════════════════════════════════════════════════════════════
# Лист 3 — Подписчики
# ═══════════════════════════════════════════════════════════════════════════════
SUB_HEADERS = ["Дата", "Подписчиков всего", "Прирост за день"]
SUB_WIDTHS  = [13, 20, 18]

def write_subs_sheet(ws, sub_intervals, is_new=True):
    if is_new:
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 36
        for ci, h in enumerate(SUB_HEADERS, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            header_style(cell)
        for ci, w in enumerate(SUB_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    existing_dates = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            existing_dates.add(str(row[0]))

    added = 0
    for idx, item in enumerate(sub_intervals):
        if str(item["date"]) in existing_dates:
            continue
        ri = ws.max_row + 1
        bg = ALT_ROW_BG if (ri % 2 == 0) else None

        prev_sub = sub_intervals[idx-1]["subscribers"] if idx > 0 else item["subscribers"]
        diff = item["subscribers"] - prev_sub

        vals = [item["date"], item["subscribers"], diff]
        fmts = [None, "#,##0", "+#,##0;-#,##0;0"]
        for ci, (v, fm) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            diff_bg = "E8F5E9" if (ci == 3 and diff > 0) else \
                      "FFEBEE" if (ci == 3 and diff < 0) else bg
            data_style(cell, bg=diff_bg, align="center" if ci==1 else "right", fmt=fm)
        added += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(SUB_HEADERS))}1"
    return added


# ═══════════════════════════════════════════════════════════════════════════════
# Лист 4 — Сводка
# ═══════════════════════════════════════════════════════════════════════════════
def write_summary_sheet(ws, channel_total, snapshot_date, pub_count):
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 14

    # Title
    title_cell = ws["A1"]
    title_cell.value = "📊 Аналитика Дзен — Сводка"
    title_cell.font = Font(name="Arial", bold=True, size=14, color=ACCENT)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws["A2"].value = f"Последнее обновление:"
    ws["A2"].font = Font(name="Arial", size=10, color="666666")
    ws["B2"].value = snapshot_date
    ws["B2"].font = Font(name="Arial", bold=True, size=10, color=ACCENT)

    ws["A3"].value = f"Всего публикаций:"
    ws["A3"].font = Font(name="Arial", size=10, color="666666")
    ws["B3"].value = pub_count
    ws["B3"].font = Font(name="Arial", bold=True, size=10, color=ACCENT)
    ws["B3"].number_format = "#,##0"

    # Section header
    ws.row_dimensions[5].height = 24
    for ci, h in enumerate(["Метрика", "Значение (всего)", "Описание"], 1):
        cell = ws.cell(row=5, column=ci)
        cell.value = h
        header_style(cell, bg=ACCENT)

    metrics = [
        ("Показы",                    channel_total["impressions"],   "Сколько раз публикации показаны"),
        ("Открытия (клики)",          channel_total["page_views"],    "Сколько раз пользователи открыли"),
        ("Дочитывания",               channel_total["deep_views"],    "Дочитали статью до конца"),
        ("Просмотры видео/гифок",     channel_total["type_views"],    "Просмотрели видео/гифку"),
        ("Новые подписчики",          channel_total["subscriptions"], "Подписались через публикации"),
        ("Лайки",                     channel_total["likes"],         "Поставили лайк"),
        ("Комментарии",               channel_total["comments"],      "Оставили комментарий"),
        ("Репосты",                   channel_total["shares"],        "Поделились публикацией"),
        ("Время просмотра, мин",      channel_total["view_time_min"], "Суммарное время вовлечённого просмотра"),
    ]

    for ri, (name, val, desc) in enumerate(metrics, 6):
        bg = ALT_ROW_BG if ri % 2 == 0 else None
        ws.row_dimensions[ri].height = 18

        c_name = ws.cell(row=ri, column=1, value=name)
        c_val  = ws.cell(row=ri, column=2, value=val)
        c_desc = ws.cell(row=ri, column=3, value=desc)

        data_style(c_name, bg=bg, align="left", bold=True)
        data_style(c_val,  bg=bg, align="right", fmt="#,##0.##")
        data_style(c_desc, bg=bg, align="left")

    # Formula references note
    note_row = 6 + len(metrics) + 1
    ws.cell(row=note_row, column=1,
            value="⚠ Данные обновляются ежедневно скриптом dzen_update.py").font = \
        Font(name="Arial", size=9, italic=True, color="999999")
    ws.cell(row=note_row+1, column=1,
            value="📂 История — на листах: Публикации | Канал_по_дням | Подписчики").font = \
        Font(name="Arial", size=9, italic=True, color=ACCENT)


# ═══════════════════════════════════════════════════════════════════════════════
# Главная функция
# ═══════════════════════════════════════════════════════════════════════════════
def main(data_json: str):
    data = json.loads(data_json)
    pubs           = data["pubs"]
    daily_intervals= data["dailyIntervals"]
    sub_intervals  = data["subIntervals"]
    channel_total  = data["channelTotal"]
    pub_count      = data.get("pubCount", len(pubs))
    snapshot_date  = datetime.now().strftime("%Y-%m-%d")

    is_new = not os.path.exists(EXCEL_PATH)

    if is_new:
        wb = Workbook()
        wb.remove(wb.active)   # remove default sheet
        ws_summary = wb.create_sheet("Сводка")
        ws_pubs    = wb.create_sheet("Публикации")
        ws_daily   = wb.create_sheet("Канал_по_дням")
        ws_subs    = wb.create_sheet("Подписчики")
    else:
        wb = load_workbook(EXCEL_PATH)
        ws_summary = wb["Сводка"]
        ws_pubs    = wb["Публикации"]
        ws_daily   = wb["Канал_по_дням"]
        ws_subs    = wb["Подписчики"]

    # Write data
    write_publications_sheet(ws_pubs, pubs, snapshot_date, is_new=is_new)
    daily_added = write_daily_sheet(ws_daily, daily_intervals, is_new=is_new)
    subs_added  = write_subs_sheet(ws_subs, sub_intervals, is_new=is_new)

    # Always refresh summary
    for row in ws_summary.iter_rows():
        for cell in row:
            cell.value = None
    write_summary_sheet(ws_summary, channel_total, snapshot_date, pub_count)

    wb.save(EXCEL_PATH)

    # Export JSON for HTML dashboard
    json_path = os.path.join(SCRIPT_DIR, "dzen_data.json")

    # ── Загружаем предыдущий снапшот для вычисления дельт pub_history ──────────
    pub_history: dict = {}
    prev_pubs_by_id: dict = {}
    try:
        if os.path.exists(json_path):
            with open(json_path, "r", encoding="utf-8") as _f:
                _old = json.load(_f)
            pub_history = _old.get("pub_history", {})
            for _p in _old.get("pubs", []):
                prev_pubs_by_id[_p["id"]] = _p
    except Exception:
        pub_history = {}
        prev_pubs_by_id = {}

    last_subs = sub_intervals[-1]["subscribers"] if sub_intervals else 0
    dashboard_pubs = [{
        "snapshot": snapshot_date,
        "id": p["id"], "title": p["title"], "type": p["type"], "pub_date": p["pub_date"],
        "impressions": p["impressions"], "opens": p["page_views"],
        "deep_views": p["deep_views"], "video_views": p["type_specific_views"],
        "subscriptions": p["subscriptions"], "likes": p["likes"],
        "comments": p["comments"], "shares": p["shares"],
        "view_time": p["view_time_min"], "ctr": p["ctr_pct"],
        "deep_rate": p["deep_views_rate_pct"]
    } for p in pubs]

    # ── Накапливаем pub_history (дельты по каждой статье) ──────────────────────
    for pub in dashboard_pubs:
        art_id = pub["id"]
        prev = prev_pubs_by_id.get(art_id, {})

        delta_impressions   = max(0, pub["impressions"]   - prev.get("impressions",   0))
        delta_opens         = max(0, pub["opens"]         - prev.get("opens",         0))
        delta_deep_views    = max(0, pub["deep_views"]    - prev.get("deep_views",    0))
        delta_likes         = max(0, pub["likes"]         - prev.get("likes",         0))
        delta_subscriptions = max(0, pub["subscriptions"] - prev.get("subscriptions", 0))

        # Не пишем запись если все дельты нулевые
        if delta_impressions == delta_opens == delta_deep_views == delta_likes == delta_subscriptions == 0:
            continue

        if art_id not in pub_history:
            pub_history[art_id] = []

        # Не дублируем запись за ту же дату
        existing_dates = {entry["date"] for entry in pub_history[art_id]}
        if snapshot_date not in existing_dates:
            pub_history[art_id].append({
                "date":          snapshot_date,
                "impressions":   delta_impressions,
                "opens":         delta_opens,
                "deep_views":    delta_deep_views,
                "likes":         delta_likes,
                "subscriptions": delta_subscriptions,
            })

    dashboard_data = {
        "updated": snapshot_date,
        "pubs": dashboard_pubs,
        "daily": [{
            "date": d["date"], "impressions": d["impressions"],
            "opens": d["page_views"], "deep_views": d["deep_views"],
            "video_views": d["type_views"], "subscriptions": d["subscriptions"],
            "likes": d["likes"], "view_time": d["view_time_min"]
        } for d in daily_intervals],
        "subs": [{
            "date": s["date"], "subscribers": s["subscribers"],
            "diff": sub_intervals[i]["subscribers"] - sub_intervals[i-1]["subscribers"] if i > 0 else 0
        } for i, s in enumerate(sub_intervals)],
        "pub_history": pub_history,
    }
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(dashboard_data, f, ensure_ascii=False, indent=2)

    # Embed fresh data into dzen_dashboard.html so it works when opened as file://
    dashboard_html = os.path.join(SCRIPT_DIR, "dzen_dashboard.html")
    if os.path.exists(dashboard_html):
        try:
            import re as _re
            with open(dashboard_html, "r", encoding="utf-8") as _f:
                _html = _f.read()
            _data_js = json.dumps(dashboard_data, ensure_ascii=False)
            _html_new = _re.sub(
                r'const EMBEDDED = \{.*?\};',
                'const EMBEDDED = ' + _data_js + ';',
                _html, flags=_re.DOTALL
            )
            if _html_new != _html:
                with open(dashboard_html, "w", encoding="utf-8") as _f:
                    _f.write(_html_new)
        except Exception:
            pass

    result = {
        "status": "ok",
        "excel_path": EXCEL_PATH,
        "json_path": json_path,
        "snapshot_date": snapshot_date,
        "pubs_written": len(pubs),
        "daily_added": daily_added,
        "subs_added": subs_added,
        "is_new_file": is_new
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return result


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python dzen_update.py '<json_data>'")
        sys.exit(1)
    main(sys.argv[1])
